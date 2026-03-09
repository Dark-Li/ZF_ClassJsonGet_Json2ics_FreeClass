from flask import Flask, render_template, request, redirect, url_for, flash, send_file
import json
import os
from datetime import datetime, timedelta
import uuid
from icalendar import Calendar
import openpyxl
from config import SEMESTER_START_DATETIME

# 导入现有的课程表获取模块
# 尝试导入course_json_out模块，如果不存在则使用模拟模块
try:
    from course_json_out import Client
except ImportError:
    # 模拟Client类，用于测试
    class Client:
        def __init__(self, base_url):
            self.base_url = base_url
            self.cookies = {}
            self.sid = None
        
        def login(self, sid, password):
            # 保存学号
            self.sid = sid
            # 模拟登录成功
            return {"code": 1000, "msg": "登录成功"}
        
        def get_info(self):
            # 模拟获取个人信息，使用学号作为姓名
            # 注意：这里应该从登录信息中获取真实姓名，这里只是模拟
            return {"code": 1000, "data": {"name": self.sid}}
        
        def get_schedule(self, year, term):
            # 模拟获取课程表
            return {
                "code": 1000,
                "data": {
                    "courses": [
                        {
                            "title": "高等数学",
                            "place": "教1-101",
                            "teacher": "张老师",
                            "credit": "4",
                            "weekday": 1,
                            "list_sessions": [1, 2],
                            "list_weeks": [1, 2, 3, 4, 5]
                        },
                        {
                            "title": "大学物理",
                            "place": "教2-201",
                            "teacher": "李老师",
                            "credit": "3",
                            "weekday": 3,
                            "list_sessions": [3, 4],
                            "list_weeks": [1, 2, 3, 4, 5]
                        }
                    ]
                }
            }


app = Flask(__name__)
app.secret_key = 'your_secret_key_here'  # 用于Flash消息

# 教务系统URL配置
BASE_URL = "https://jxw.sylu.edu.cn"  # 替换为实际的教务系统URL

# 目录配置
JSON_FOLDER = "json_input"
ICS_FOLDER = "ics_output"
EXCEL_FOLDER = "空课表统计结果Excel"

# 学期开始日期
SEMESTER_START = SEMESTER_START_DATETIME

# 每节课对应的时间
SESSION_TIMES = {
    1: ("08:00", "08:45"),
    2: ("08:55", "09:40"),
    3: ("10:00", "10:45"),
    4: ("10:55", "11:40"),
    5: ("13:00", "13:45"),
    6: ("13:55", "14:40"),
    7: ("14:50", "15:35"),
    8: ("15:45", "16:30"),
    9: ("16:40", "17:25"),
    10: ("17:35", "18:20"),
    11: ("18:30", "19:15"),
    12: ("19:25", "20:10"),
}

# 课程时间段定义
CLASS_TIME_SLOTS = [
    ("第1-2节", (8, 0), (9, 40)),
    ("第3-4节", (10, 0), (11, 40)),
    ("第5-6节", (13, 0), (14, 40)),
    ("第7-8节", (14, 50), (16, 30)),
    ("第9-10节", (16, 40), (18, 20)),
    ("第11-12节", (18, 30), (20, 10))
]

# 确保目录存在
os.makedirs(JSON_FOLDER, exist_ok=True)
os.makedirs(ICS_FOLDER, exist_ok=True)
os.makedirs(EXCEL_FOLDER, exist_ok=True)

# 主页面
@app.route('/')
def index():
    return render_template('index.html')

# ==================== GetJsonClass 功能 ====================
@app.route('/get_json_class')
def get_json_class():
    return render_template('GetJsonClass.html')

@app.route('/get_schedule', methods=['POST'])
def get_schedule():
    sid = request.form['sid']
    password = request.form['password']
    
    if not sid or not password:
        flash('请输入完整的学号和密码', 'error')
        return redirect(url_for('get_json_class'))
    
    try:
        # 创建客户端实例
        lgn = Client(base_url=BASE_URL)
        
        # 登录
        pre_login = lgn.login(sid, password)
        
        if pre_login["code"] == 1001:
            # 需要验证码，目前不支持，返回错误
            flash('登录需要验证码，请使用其他方式获取课程表', 'error')
            return redirect(url_for('get_json_class'))
        elif pre_login["code"] == 1000:
            # 登录成功
            lgn_cookies = lgn.cookies
        else:
            # 登录失败
            flash(f'登录失败：{pre_login["msg"]}', 'error')
            return redirect(url_for('get_json_class'))
        
        # 获取个人信息
        self_info = lgn.get_info()
        if self_info["code"] != 1000:
            flash(f'获取个人信息失败：{self_info["msg"]}', 'error')
            return redirect(url_for('get_json_class'))
        
        name = self_info["data"]["name"]
        
        # 获取当前学年学期
        current_year = datetime.now().year
        current_month = datetime.now().month
        # 8-12月为第一学期，1-7月为第二学期
        current_term = 1 if 8 <= current_month <= 12 else 2
        # 学年为当前年份（第一学期）或当前年份-1（第二学期）
        academic_year = current_year if current_term == 1 else current_year - 1
        
        # 获取课程表
        course_json = lgn.get_schedule(academic_year, current_term)
        if course_json["code"] != 1000:
            flash(f'获取课程表失败：{course_json["msg"]}', 'error')
            return redirect(url_for('get_json_class'))
        
        # 保存JSON文件
        file_path = os.path.join(JSON_FOLDER, f"{name}.json")
        with open(file_path, "w", encoding="utf-8") as file:
            file.write(json.dumps(course_json, ensure_ascii=False, indent=4))
        
        flash(f'课程表已成功获取并保存，姓名：{name}', 'success')
        return redirect(url_for('get_json_class'))
        
    except Exception as e:
        flash(f'获取课程表时发生错误：{str(e)}', 'error')
        return redirect(url_for('get_json_class'))

# ==================== Json2ics 和统计空课表 功能 ====================
@app.route('/json2ics')
def json2ics_page():
    # 获取json_input和ics_output目录中的文件列表
    json_files = [f for f in os.listdir(JSON_FOLDER) if f.endswith('.json')]
    ics_files = [f for f in os.listdir(ICS_FOLDER) if f.endswith('.ics')]
    excel_files = [f for f in os.listdir(EXCEL_FOLDER) if f.endswith('.xlsx')]
    return render_template('json2ics.html', json_files=json_files, ics_files=ics_files, excel_files=excel_files)

# 工具函数
def get_datetime_for_course(week, weekday, session):
    """根据周数、星期、节数计算具体日期和时间"""
    day = SEMESTER_START + timedelta(weeks=week-1, days=weekday-1)
    start_time_str, end_time_str = SESSION_TIMES[session]
    start_dt = datetime.strptime(f"{day.date()} {start_time_str}", "%Y-%m-%d %H:%M")
    end_dt = datetime.strptime(f"{day.date()} {end_time_str}", "%Y-%m-%d %H:%M")
    return start_dt, end_dt

def format_datetime(dt):
    """格式化为ICS标准时间格式"""
    return dt.strftime("%Y%m%dT%H%M%S")

def generate_uid():
    """生成唯一UID"""
    return uuid.uuid4().hex

def generate_ics_from_json(json_path, ics_path):
    """根据单个 JSON 文件生成 ICS 文件"""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    courses = data["data"]["courses"]

    ics_content = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//Kagg886//ICS Generator//EN",
        "CALSCALE:GREGORIAN"
    ]

    for course in courses:
        title = course["title"]
        location = course["place"]
        teacher = course["teacher"]
        credit = course.get("credit", "")
        weekday = course["weekday"]  # 1=周一
        list_sessions = course["list_sessions"]
        list_weeks = course["list_weeks"]

        for week in list_weeks:
            for session in list_sessions:
                start_dt, end_dt = get_datetime_for_course(week, weekday, session)
                dtstamp = datetime.now().strftime("%Y%m%dT%H%M%SZ")
                uid = generate_uid()
                description = f"1. 任课教师: {teacher}\n2. 课程属性: 必修\n3. 学分: {credit}\n4. 属于系统课程: 否"

                ics_content.extend([
                    "BEGIN:VEVENT",
                    f"UID:{uid}",
                    f"DTSTAMP:{dtstamp}",
                    f"DTSTART;TZID=Asia/Shanghai:{format_datetime(start_dt)}",
                    f"DTEND;TZID=Asia/Shanghai:{format_datetime(end_dt)}",
                    f"SUMMARY:{title}",
                    f"DESCRIPTION:{description}",
                    f"LOCATION:{location}",
                    "BEGIN:VALARM",
                    "ACTION:AUDIO",
                    "TRIGGER:-PT30M",
                    f"DESCRIPTION:{title} 即将开始",
                    "END:VALARM",
                    "END:VEVENT"
                ])

    ics_content.append("END:VCALENDAR")

    # 写入 ICS 文件
    with open(ics_path, "w", encoding="utf-8") as f:
        f.write("\n".join(ics_content))

# 解析单个 .ics 文件
def parse_ics(file_path, person_name):
    from datetime import time
    busy = {}  # week -> day -> [(start_time, end_time), ...]

    with open(file_path, "r", encoding="utf-8") as f:
        cal = Calendar.from_ical(f.read())

    for component in cal.walk():
        if component.name != "VEVENT":
            continue

        dtstart = component.get("DTSTART").dt
        dtend = component.get("DTEND").dt

        # 跳过无效数据
        if not isinstance(dtstart, datetime):
            continue

        date = dtstart.date()
        start_time = dtstart.time()
        end_time = dtend.time()

        # 计算周次与星期
        week = ((date - SEMESTER_START.date()).days // 7) + 1
        weekday = date.isoweekday()  # 1=周一, 7=周日

        if week < 1 or week > 20:
            continue

        busy.setdefault(week, {}).setdefault(weekday, []).append((start_time, end_time))

    return busy

# 检查时间是否重叠
def is_overlap(slot_start, slot_end, class_start, class_end):
    """判断课程时间与节次时间是否重叠"""
    return not (class_end <= slot_start or class_start >= slot_end)

@app.route('/convert_json2ics', methods=['POST'])
def convert_json2ics():
    try:
        # 批量处理文件夹
        for filename in os.listdir(JSON_FOLDER):
            if filename.lower().endswith(".json"):
                json_path = os.path.join(JSON_FOLDER, filename)
                ics_filename = os.path.splitext(filename)[0] + ".ics"
                ics_path = os.path.join(ICS_FOLDER, ics_filename)
                generate_ics_from_json(json_path, ics_path)
        
        flash('所有JSON文件已成功转换为ICS文件', 'success')
    except Exception as e:
        flash(f'转换过程中发生错误：{str(e)}', 'error')
    
    return redirect(url_for('json2ics_page'))

@app.route('/generate_empty_course_stat', methods=['POST'])
def generate_empty_course_stat():
    try:
        from datetime import time
        # 批量解析
        all_busy = {}
        for filename in os.listdir(ICS_FOLDER):
            if filename.lower().endswith(".ics"):
                name = os.path.splitext(filename)[0]
                path = os.path.join(ICS_FOLDER, filename)
                all_busy[name] = parse_ics(path, name)
        
        # 生成 Excel
        wb = openpyxl.Workbook()

        for week in range(1, 21):
            ws = wb.create_sheet(f"第{week}周")
            headers = ["时间段", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]
            ws.append(headers)

            for slot_name, (sh, sm), (eh, em) in CLASS_TIME_SLOTS:
                slot_start = time(sh, sm)
                slot_end = time(eh, em)
                row = [slot_name]

                for weekday in range(1, 8):
                    free_people = []

                    for person, busy_dict in all_busy.items():
                        busy_times = busy_dict.get(week, {}).get(weekday, [])
                        # 检查是否有课与该节次重叠
                        has_class = any(is_overlap(slot_start, slot_end, s, e) for s, e in busy_times)
                        if not has_class:
                            free_people.append(person)

                    row.append(", ".join(free_people))

                ws.append(row)

        # 删除默认Sheet
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        output_file = os.path.join(EXCEL_FOLDER, f"空课统计表_增强版_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
        wb.save(output_file)
        
        flash(f'空课统计表已成功生成：{os.path.basename(output_file)}', 'success')
    except Exception as e:
        flash(f'生成统计表格时发生错误：{str(e)}', 'error')
    
    return redirect(url_for('json2ics_page'))

@app.route('/download/<folder>/<filename>')
def download_file(folder, filename):
    # 安全检查，确保只能下载指定目录的文件
    if folder == 'ics':
        file_path = os.path.join(ICS_FOLDER, filename)
    elif folder == 'excel':
        file_path = os.path.join(EXCEL_FOLDER, filename)
    else:
        return "访问被拒绝", 403
    
    if not os.path.exists(file_path):
        return "文件不存在", 404
    
    return send_file(file_path, as_attachment=True)

# ==================== WebView 功能 ====================
# 导入用户写好的ZF_ClassWebView模块
import importlib.util
import sys

# 加载ZF_ClassWebView.py模块
spec = importlib.util.spec_from_file_location("ZF_ClassWebView", "ZF_ClassWebView.py")
ZF_ClassWebView = importlib.util.module_from_spec(spec)
sys.modules["ZF_ClassWebView"] = ZF_ClassWebView
spec.loader.exec_module(ZF_ClassWebView)

# 使用ZF_ClassWebView中的index函数作为web_view路由的处理函数
@app.route('/web_view', methods=['GET', 'POST'])
def web_view():
    # 调用ZF_ClassWebView中的index函数
    return ZF_ClassWebView.index()

# 导入必要的模块
from datetime import timedelta

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)