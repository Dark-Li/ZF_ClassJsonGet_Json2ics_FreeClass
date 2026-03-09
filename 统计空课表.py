import os
import datetime
from icalendar import Calendar
from openpyxl import Workbook
from config import SEMESTER_START_DATE

# ==== 基本配置 ====
START_DATE = SEMESTER_START_DATE  # 第1周周一
ICS_FOLDER = "./ics_output"                # ics 文件所在文件夹
OUTPUT_FILE = "./空课表统计结果Excel/空课统计表_增强版.xlsx"

# ==== 节次时间表 ====
TIME_SLOTS = [
    ("第1-2节", (8, 0), (9, 40)),
    ("第3-4节", (10, 0), (11, 40)),
    ("第5-6节", (13, 0), (14, 40)),
    ("第7-8节", (14, 50), (16, 30)),
    ("第9-10节", (16, 40), (18, 20)),
    ("第11-12节", (18, 30), (20, 10)),
]

# ==== 解析单个 .ics 文件 ====
def parse_ics(file_path, person_name):
    busy = {}  # week -> day -> [(start_time, end_time), ...]

    with open(file_path, "r", encoding="utf-8") as f:
        cal = Calendar.from_ical(f.read())

    for component in cal.walk():
        if component.name != "VEVENT":
            continue

        dtstart = component.get("DTSTART").dt
        dtend = component.get("DTEND").dt

        # 跳过无效数据
        if not isinstance(dtstart, datetime.datetime):
            continue

        date = dtstart.date()
        start_time = dtstart.time()
        end_time = dtend.time()

        # 计算周次与星期
        week = ((date - START_DATE).days // 7) + 1
        weekday = date.isoweekday()  # 1=周一, 7=周日

        if week < 1 or week > 20:
            continue

        busy.setdefault(week, {}).setdefault(weekday, []).append((start_time, end_time))

    return busy

# ==== 检查时间是否重叠 ====
def is_overlap(slot_start, slot_end, class_start, class_end):
    """判断课程时间与节次时间是否重叠"""
    return not (class_end <= slot_start or class_start >= slot_end)

# ==== 批量解析 ====
all_busy = {}
for filename in os.listdir(ICS_FOLDER):
    if filename.lower().endswith(".ics"):
        name = os.path.splitext(filename)[0]
        path = os.path.join(ICS_FOLDER, filename)
        all_busy[name] = parse_ics(path, name)
        print(f"✅ 解析课表: {name}")

# ==== 生成 Excel ====
wb = Workbook()

for week in range(1, 21):
    ws = wb.create_sheet(f"第{week}周")
    headers = ["时间段", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    ws.append(headers)

    for slot_name, (sh, sm), (eh, em) in TIME_SLOTS:
        slot_start = datetime.time(sh, sm)
        slot_end = datetime.time(eh, em)
        row = [slot_name]

        for weekday in range(1, 8):
            free_people = []

            for person, busy_dict in all_busy.items():
                busy_times = busy_dict.get(week, {}).get(weekday, [])
                # 检查是否有课与该节次重叠
                has_class = any(is_overlap(slot_start, slot_end, s, e) for s, e in busy_times)
                if not has_class:
                    free_people.append(person)

            row.append(", ".join(free_people))

        ws.append(row)

# 删除默认Sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

wb.save(OUTPUT_FILE)
print(f"\n🎉 已生成空课统计表: {OUTPUT_FILE}")
