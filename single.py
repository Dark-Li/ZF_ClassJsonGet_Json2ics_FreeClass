import os
import datetime
from icalendar import Calendar
from openpyxl import Workbook
from config import SEMESTER_START_DATE

# ==== 参数设置 ====
START_DATE = SEMESTER_START_DATE  # 第1周周一
PERSON_NAME = "郑佳琦"
ICS_FILE = "ics_output/郑佳琦.ics"
OUTPUT_FILE = "空课统计表_单人测试.xlsx"

# 节次时间段（起始、结束）
TIME_SLOTS = [
    ("第1-2节", (8, 0), (9, 40)),
    ("第3-4节", (10, 0), (11, 40)),
    ("第5-6节", (13, 0), (14, 40)),
    ("第7-8节", (14, 50), (16, 30)),
    ("第9-10节", (16, 40), (18, 20)),
    ("第11-12节", (18, 30), (20, 10)),
]

# ==== 读取 .ics ====
with open(ICS_FILE, "r", encoding="utf-8") as f:
    cal = Calendar.from_ical(f.read())

# 存储该人所有课程时间
busy_times = {}  # week -> day -> list of time ranges

for component in cal.walk():
    if component.name == "VEVENT":
        dtstart = component.get("DTSTART").dt
        dtend = component.get("DTEND").dt

        # 转换为日期和时间
        if isinstance(dtstart, datetime.datetime):
            date = dtstart.date()
            start_time = dtstart.time()
        else:
            continue  # 忽略无时间的事件

        # 计算星期（1=周一 ... 7=周日）
        weekday = date.isoweekday()

        # 计算第几周
        week = ((date - START_DATE).days // 7) + 1
        if week < 1 or week > 20:
            continue

        busy_times.setdefault(week, {}).setdefault(weekday, []).append((start_time, dtend.time()))

# ==== 判断每节课是否空闲 ====
wb = Workbook()
for week in range(1, 21):
    ws = wb.create_sheet(f"第{week}周")

    # 表头
    headers = ["时间段", "周一", "周二", "周三", "周四", "周五", "周六", "周日"]
    ws.append(headers)

    for slot_name, (start_h, start_m), (end_h, end_m) in TIME_SLOTS:
        row = [slot_name]
        slot_start = datetime.time(start_h, start_m)
        slot_end = datetime.time(end_h, end_m)

        for weekday in range(1, 8):
            busy = False
            for s, e in busy_times.get(week, {}).get(weekday, []):
                if not (e <= slot_start or s >= slot_end):
                    busy = True
                    break
            if not busy:
                row.append(PERSON_NAME)
            else:
                row.append("")
        ws.append(row)

# 删除默认Sheet
if "Sheet" in wb.sheetnames:
    del wb["Sheet"]

wb.save(OUTPUT_FILE)
print(f"✅ 已生成：{OUTPUT_FILE}")
